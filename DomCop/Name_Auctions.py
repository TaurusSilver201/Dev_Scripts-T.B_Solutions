import os
import glob
import shutil
import config
import ftplib
import time
import openpyxl
import random
import datetime
import pandas as pd
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select


URLS_G = ["https://www.domcop.com/domains?sid=19381",
    "https://www.domcop.com/domains?sid=19382",
    "https://www.domcop.com/domains?sid=19384",
    "https://www.domcop.com/domains?sid=19383"]
    
URLS_NJ = ["https://www.domcop.com/domains?sid=19387",
    "https://www.domcop.com/domains?sid=19388",
    "https://www.domcop.com/domains?sid=19389",
    "https://www.domcop.com/domains?sid=19390"]
    
URLS_A2 = ["https://www.domcop.com/domains?sid=19401",
    "https://www.domcop.com/domains?sid=19402",
    "https://www.domcop.com/domains?sid=19403"]

    
    
service = webdriver.ChromeService()
options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')
options.add_argument("--disable-dev-shm-usage")
options.add_argument('log-level=3')
options.add_argument(r'user-data-dir=C:\Users\User3\AppData\Local\Google\Chrome\User Data')
prefs = {"download.default_directory" : r"C:\DomCop\Name\Files"}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(service=service, options=options)
print('.: NAME_AUCTIONS.BAT :.')

current_date = datetime.date.today()

exp_date = config.exp_date_today.strftime("%d-%m-%Y")
exp_date3 = (current_date + datetime.timedelta(days=3)).strftime("%d-%m-%Y")
exp_date4 = (current_date + datetime.timedelta(days=4)).strftime("%d-%m-%Y")
exp_date5 = (current_date + datetime.timedelta(days=5)).strftime("%d-%m-%Y")
date_exp_3_days = (current_date + datetime.timedelta(days=3)).strftime("%Y-%m-%d")
date_exp_4_days = (current_date + datetime.timedelta(days=4)).strftime("%Y-%m-%d")
date_exp_5_days = (current_date + datetime.timedelta(days=5)).strftime("%Y-%m-%d")

def split_csv_by_date(input_file, output_prefix):
    df = pd.read_csv(input_file)

    # Ensure consistent date format (YYYY-MM-DD) for grouping
    try:
        df['Exp Date (UTC)'] = pd.to_datetime(df['Exp Date (UTC)'], utc=True).dt.strftime('%d-%m-%Y')
        grouped_data = df.groupby('Exp Date (UTC)')
    except:
        df['Exp Date (UTC)'] = pd.to_datetime(df['Exp Date (UTC)'], format='mixed').dt.strftime('%d-%m-%Y')
        grouped_data = df.groupby('Exp Date (UTC)')

    # Split all files first
    for date, group_df in grouped_data:
        output_file = f"{output_prefix}_{date}.csv"
        group_df.to_csv(output_file, index=False)
    
    print('file splitted..')

    
# Function to get the date from the filename
def get_date_from_filename(filename):
    parts = filename.split('_')
    if len(parts) > 1:
        try:
            date_str = parts[-1].split('.')[0]
            the_date = datetime.datetime.strptime(date_str, '%d-%m-%Y').date()
            return the_date
        except ValueError:
            pass
    return None

# Function to check if a date is within +3, +4, or +5 days from today
def is_within_3_5_days(date):
    today = datetime.date.today()
    for i in range(3, 6):
        target_date = today + datetime.timedelta(days=i)
        if date == target_date:
            return True
    return False


def filter_bids_and_price(csv_file, excel_file):
    # Load your CSV file into a DataFrame
    df = pd.read_csv(csv_file, header=0, index_col=False)

    # Convert 'Bids' column to numeric type
    df['Bids'] = pd.to_numeric(df['Bids'], errors='coerce')

    # Strip leading and trailing whitespaces from 'Price' column, remove '$', then convert to numeric type
    df['Price'] = pd.to_numeric(df['Price'].str.strip().replace('[\$,]', '', regex=True), errors='coerce')

    # Apply filtering conditions
    filtered_df = df[(df['Bids'] > 3) | (df['Price'] < 149)]

     # Add dollar sign ('$') back to the 'Price' column
    filtered_df['Price'] = filtered_df['Price'].apply(lambda x: f"${x}")

    # Write the filtered DataFrame to a new CSV file
    filtered_df.to_csv(excel_file + ".csv", index=False)



def random_sleep(x, y):
    sleep_time = random.uniform(x, y)
    time.sleep(sleep_time)
    
def expireddomains1(date_exp, date_name):
    expireddomainsurl = 'https://member.expireddomains.net/savedsearches/goto/429443/'
    driver.get(expireddomainsurl)
    print('opening expireddomain..')
    random_sleep(5, 10)

    domain_data_list = []  # List to store extracted domain data

    try:
        # Find the "Show Filter" button (assuming it's always present)
        show_filter_button = driver.find_element(By.LINK_TEXT, "Show Filter")
        print('show filter..')

        # Click the button
        show_filter_button.click()
        random_sleep(1, 3)

        # Select the date by ID
        select = Select(driver.find_element(By.ID, 'fenddate'))
        select.select_by_value(date_exp)
        print('choose date..')
        random_sleep(1, 3)

        # Find the "Submit" button by name
        button = driver.find_element(By.NAME, "button_submit")
        
        # Click the "Submit" button and wait
        button.click()
        print('apply filter..')
        random_sleep(3, 10)

        clickable = True
            
        while clickable == True:
            # Find domain data elements
            span_element = driver.find_element(By.CSS_SELECTOR, 'span[data-clipboard-text]')
            data_text = span_element.get_attribute('data-clipboard-text')
            domain_data_list.append(data_text)
            print('domain copied into list..')
            random_sleep(3, 5)
            
            next_button = ''
            try:
                #Check for "Next" button and click if available
                next_buttons = driver.find_elements(By.CLASS_NAME, "next")
                next_button = next_buttons[0]
                if next_button:
                    next_button.click()
                    print('next page..')
                    random_sleep(10, 15)
            except:
                clickable = False
                # No "Next" button found, break the loop
                break
               
        # Create the filename and folder structure
        folder_path = r"C:\DomCop\Name\Files"
        filename = f"Name_GD_Expireddomains_{date_name}.txt"
        os.makedirs(folder_path, exist_ok=True)

        # Construct the full filepath
        filepath = os.path.join(folder_path, filename)

        # Open the file in write mode and write all extracted data at once
        with open(filepath, "w") as text_file:
            text_file.write("\n".join(domain_data_list))  # Efficient single write

        # Go to Google (optional)
        driver.get('https://google.com')

    except Exception as e:
        print(f"Something went wrong: {e}")
        driver.quit()


def expireddomains2(date_exp, date_name):
    expireddomainsurl = 'https://member.expireddomains.net/savedsearches/goto/429445/'
    driver.get(expireddomainsurl)
    print('opening expireddomain..')
    random_sleep(5, 10)

    domain_data_list = []  # List to store extracted domain data

    try:
        # Find the "Show Filter" button (assuming it's always present)
        show_filter_button = driver.find_element(By.LINK_TEXT, "Show Filter")
        print('show filter..')

        # Click the button
        show_filter_button.click()
        random_sleep(1, 3)

        # Select the date by ID
        select = Select(driver.find_element(By.ID, 'fenddate'))
        select.select_by_value(date_exp)
        print('choose date..')
        random_sleep(1, 3)

        # Find the "Submit" button by name
        button = driver.find_element(By.NAME, "button_submit")
        
        # Click the "Submit" button and wait
        button.click()
        print('apply filter..')
        random_sleep(3, 10)

        clickable = True
            
        while clickable == True:
            # Find domain data elements
            span_element = driver.find_element(By.CSS_SELECTOR, 'span[data-clipboard-text]')
            data_text = span_element.get_attribute('data-clipboard-text')
            domain_data_list.append(data_text)
            print('domain copied into list..')
            random_sleep(3, 5)
            
            next_button = ''
            try:
                #Check for "Next" button and click if available
                next_buttons = driver.find_elements(By.CLASS_NAME, "next")
                next_button = next_buttons[0]
                if next_button:
                    next_button.click()
                    print('next page..')
                    random_sleep(10, 15)
            except:
                clickable = False
                # No "Next" button found, break the loop
                break
               
        # Create the filename and folder structure
        folder_path = r"C:\DomCop\Name\Files"
        filename = f"Name_NJ_Expireddomains_{date_name}.txt"
        os.makedirs(folder_path, exist_ok=True)

        # Construct the full filepath
        filepath = os.path.join(folder_path, filename)

        # Open the file in write mode and write all extracted data at once
        with open(filepath, "w") as text_file:
            text_file.write("\n".join(domain_data_list))  # Efficient single write

        # Go to Google (optional)
        driver.get('https://google.com')

    except Exception as e:
        print(f"Something went wrong: {e}")
        driver.quit()
        

def expireddomains3(date_exp, date_name):
    expireddomainsurl = 'https://member.expireddomains.net/savedsearches/goto/429450/'
    driver.get(expireddomainsurl)
    print('opening expireddomain..')
    random_sleep(5, 10)

    domain_data_list = []  # List to store extracted domain data

    try:
        # Find the "Show Filter" button (assuming it's always present)
        show_filter_button = driver.find_element(By.LINK_TEXT, "Show Filter")
        print('show filter..')

        # Click the button
        show_filter_button.click()
        random_sleep(1, 3)

        # Select the date by ID
        select = Select(driver.find_element(By.ID, 'fenddate'))
        select.select_by_value(date_exp)
        print('choose date..')
        random_sleep(1, 3)

        # Find the "Submit" button by name
        button = driver.find_element(By.NAME, "button_submit")
        
        # Click the "Submit" button and wait
        button.click()
        print('apply filter..')
        random_sleep(3, 10)

        clickable = True
            
        while clickable == True:
            # Find domain data elements
            span_element = driver.find_element(By.CSS_SELECTOR, 'span[data-clipboard-text]')
            data_text = span_element.get_attribute('data-clipboard-text')
            domain_data_list.append(data_text)
            print('domain copied into list..')
            random_sleep(3, 5)
            
            next_button = ''
            try:
                #Check for "Next" button and click if available
                next_buttons = driver.find_elements(By.CLASS_NAME, "next")
                next_button = next_buttons[0]
                if next_button:
                    next_button.click()
                    print('next page..')
                    random_sleep(10, 15)
            except:
                clickable = False
                # No "Next" button found, break the loop
                break
               
        # Create the filename and folder structure
        folder_path = r"C:\DomCop\Name\Files"
        filename = f"Name_Auctions2_Expireddomains_{date_name}.txt"
        os.makedirs(folder_path, exist_ok=True)

        # Construct the full filepath
        filepath = os.path.join(folder_path, filename)

        # Open the file in write mode and write all extracted data at once
        with open(filepath, "w") as text_file:
            text_file.write("\n".join(domain_data_list))  # Efficient single write

        # Go to Google (optional)
        driver.get('https://google.com')

    except Exception as e:
        print(f"Something went wrong: {e}")
        driver.quit()
        

def GD(driver, url):
    try:
        #access the web page
        driver.get(url)
        random_sleep(8, 15)
        print("accessing web")
        driver.find_element(By.XPATH, '//button[@class="btn btn-primary btn-minier"][3]').click()
        print("export..")
        random_sleep(1, 2)
        driver.find_element(By.LINK_TEXT, 'Select None').click()
        random_sleep(3, 5)
        print("choosing menu..")
        driver.find_element(By.XPATH, '//input[@name="DomainData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//button[@class="btn btn-info"][1]').click()
        random_sleep(2, 5)
        print("waiting to download file...")

        try:
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")
        except:
            print('loading more... wait!')
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")

        # link = None
        # while not link:
        #     try:
        #         time.sleep(20)
        #         link = driver.find_element(By.XPATH, '//tr[2]/td[8]/a[@href]').click()
        #         link = 'done'
        #         print("download success...")
        #     except NoSuchElementException:
        #         time.sleep(30)  
        random_sleep(2, 5)
        # Renaming the and split file
        download_path = glob.glob(r"C:\DomCop\Name\Files\*") 
        latest_filename = max(download_path, key=os.path.getctime)
        output_prefix = latest_filename[:-4]
        filter_bids_and_price(latest_filename, output_prefix)
        print('bad domain removed..')
        split_csv_by_date(latest_filename, output_prefix)
        directory = r"C:\DomCop\Name\Files"
       
        # Iterate over files in the directory
        for filename in os.listdir(directory):
            if filename.endswith(".csv"):
                file_date = get_date_from_filename(filename)
                if file_date:
                    if is_within_3_5_days(file_date) and filename.startswith("DomCop-Name-GD"):
                        pass
                    else:
                        print(f"Deleting file: {filename}")
                        os.remove(os.path.join(directory, filename))

    except Exception as error:
        print(error)
        print('something wrong.. browser quit')
        driver.quit()
        
def NJ(driver, url):
    try:
        #access the web page
        driver.get(url)
        random_sleep(8, 15)
        print("accessing web")
        driver.find_element(By.XPATH, '//button[@class="btn btn-primary btn-minier"][3]').click()
        print("export..")
        random_sleep(1, 2)
        driver.find_element(By.LINK_TEXT, 'Select None').click()
        random_sleep(3, 5)
        print("choosing menu..")
        driver.find_element(By.XPATH, '//input[@name="DomainData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//button[@class="btn btn-info"][1]').click()
        random_sleep(2, 5)
        print("waiting to download file...")

        try:
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")
        except:
            print('loading more... wait!')
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")

        # link = None
        # while not link:
        #     try:
        #         time.sleep(20)
        #         link = driver.find_element(By.XPATH, '//tr[2]/td[8]/a[@href]').click()
        #         link = 'done'
        #         print("download success...")
        #     except NoSuchElementException:
        #         time.sleep(30)  
        random_sleep(2, 5)
        # Renaming the and split file
        download_path = glob.glob(r"C:\DomCop\Name\Files\*") 
        latest_filename = max(download_path, key=os.path.getctime)
        output_prefix = latest_filename[:-4]
        filter_bids_and_price(latest_filename, output_prefix)
        print('bad domain removed..')
        split_csv_by_date(latest_filename, output_prefix)
        directory = r"C:\DomCop\Name\Files"
       
        # Iterate over files in the directory
        for filename in os.listdir(directory):
            if filename.endswith(".csv"):
                file_date = get_date_from_filename(filename)
                if file_date:
                    if is_within_3_5_days(file_date) and filename.startswith("DomCop-Name-NJ"):
                        pass
                    else:
                        print(f"Deleting file: {filename}")
                        os.remove(os.path.join(directory, filename))

    except Exception as error:
        print(error)
        print('something wrong.. browser quit')
        driver.quit()
        
def A2(driver, url):
    try:
        #access the web page
        driver.get(url)
        random_sleep(8, 15)
        print("accessing web")
        driver.find_element(By.XPATH, '//button[@class="btn btn-primary btn-minier"][3]').click()
        print("export..")
        random_sleep(1, 2)
        driver.find_element(By.LINK_TEXT, 'Select None').click()
        random_sleep(3, 5)
        print("choosing menu..")
        driver.find_element(By.XPATH, '//input[@name="DomainData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//button[@class="btn btn-info"][1]').click()
        random_sleep(2, 5)
        print("waiting to download file...")

        try:
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")
        except:
            print('loading more... wait!')
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")

        # link = None
        # while not link:
        #     try:
        #         time.sleep(20)
        #         link = driver.find_element(By.XPATH, '//tr[2]/td[8]/a[@href]').click()
        #         link = 'done'
        #         print("download success...")
        #     except NoSuchElementException:
        #         time.sleep(30)  
        random_sleep(2, 5)
        # Renaming the and split file
        download_path = glob.glob(r"C:\DomCop\Name\Files\*") 
        latest_filename = max(download_path, key=os.path.getctime)
        output_prefix = latest_filename[:-4]
        filter_bids_and_price(latest_filename, output_prefix)
        print('bad domain removed..')
        split_csv_by_date(latest_filename, output_prefix)
        directory = r"C:\DomCop\Name\Files"
       
        # Iterate over files in the directory
        for filename in os.listdir(directory):
            if filename.endswith(".csv"):
                file_date = get_date_from_filename(filename)
                if file_date:
                    if is_within_3_5_days(file_date) and filename.startswith("DomCop-Name-Auctions2"):
                        pass
                    else:
                        print(f"Deleting file: {filename}")
                        os.remove(os.path.join(directory, filename))

    except Exception as error:
        print(error)
        print('something wrong.. browser quit')
        driver.quit()
        
  
def copy_template():
    original_file = f'C:\DomCop\Template_Name_DomCop.xlsx'
    destination_directory = r'C:\DomCop\Name'
    
    # Construct the new filename
    filename1 = f"Name_GD_{exp_date3}.xlsx"
    filename2 = f"Name_GD_{exp_date4}.xlsx"
    filename3 = f"Name_GD_{exp_date5}.xlsx"
    filename4 = f"Name_NJ_{exp_date3}.xlsx"
    filename5 = f"Name_NJ_{exp_date4}.xlsx"
    filename6 = f"Name_NJ_{exp_date5}.xlsx"
    filename7 = f"Name_Auctions2_{exp_date3}.xlsx"
    filename8 = f"Name_Auctions2_{exp_date4}.xlsx"
    filename9 = f"Name_Auctions2_{exp_date5}.xlsx"
    destination_path = os.path.join(destination_directory, filename1)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename2)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename3)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename4)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename5)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename6)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename7)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename8)
    shutil.copyfile(original_file, destination_path)
    
    destination_path = os.path.join(destination_directory, filename9)
    shutil.copyfile(original_file, destination_path)

    print(f"template copied..")
    

def paste_data(file_type, txt_file, excel_file):

    domain_list_file = f"C:\\DomCop\\Name\\Files\\Name_{file_type}_Expireddomains_{txt_file}.txt"
       
    domain_list = [] 
    with open(domain_list_file, "r") as f:
        domain_list = f.readlines()
        
        
    # Clean the domain list (optional, remove leading/trailing whitespaces)
    domain_list = [domain.strip() for domain in domain_list]
        
    csv_files = []
    
    directory = r"C:\DomCop\Name\Files"    
    # Iterate over files in the directory
    temp_files = []
    for filename in os.listdir(directory):
        if filename.endswith(".csv"):
           temp_files.append(filename)
           datefix = txt_file
           csv_files = [file for file in temp_files if datefix in file]      
    
    for csv_file in csv_files:
        try:
            df = pd.read_csv(directory+'\\'+csv_file)
            domain_list.extend(df["Domain Name"].tolist())  # Efficiently append domain names
        except FileNotFoundError:
            print(f"Error: CSV file '{csv_file}' not found. Skipping...")
            
    
    # Lowercase all characters in the list
    domain_list = [domain.lower() for domain in domain_list]
    
    # Remove duplicates from domain_list using set conversion
    domain_list = list(set(domain_list))

    # Load the existing Excel file
    workbook = openpyxl.load_workbook(excel_file)

    # Check if the "filter" sheet exists
    if "filter" not in workbook.sheetnames:
        # Create the "filter" sheet at the beginning
        worksheet = workbook.create_sheet("filter", 0)
    else:
        # Get the existing "filter" sheet
        worksheet = workbook[ "filter"]

    # Clear formatting on the entire sheet
    for row in worksheet.iter_rows():
      row[1].alignment = None  # Clear alignment for the first cell (assuming data starts from A2)
      worksheet.row_dimensions[row[1].row].alignment = None 

    # Write the domain list starting from the first row (checks for existing data)
    if worksheet["A1"].value is None:
        row_counter = 1  # Start writing from the first row if empty
    else:
        row_counter = worksheet.max_row + 1  # Start after existing data
    
    row_counter = 2
    for domain in domain_list:
        worksheet.cell(row=row_counter, column=1, value=domain)
        row_counter += 1

    # Save the Excel file
    workbook.save(excel_file)
    print("Domain list pasted successfully!")


        
if __name__ == "__main__": 
    for url in URLS_G:
        GD(driver, url)
        
    expireddomains1(date_exp_3_days, exp_date3)
    expireddomains1(date_exp_4_days, exp_date4) 
    expireddomains1(date_exp_5_days, exp_date5)
    
    copy_template()
    excel_file1 = f"C:\\DomCop\\Name\\Name_GD_{exp_date3}.xlsx"
    excel_file2 = f"C:\\DomCop\\Name\\Name_GD_{exp_date4}.xlsx"
    excel_file3 = f"C:\\DomCop\\Name\\Name_GD_{exp_date5}.xlsx"
    file_type = 'GD'
    paste_data(file_type, exp_date3, excel_file1)
    paste_data(file_type, exp_date4, excel_file2)
    paste_data(file_type, exp_date5, excel_file3)
    
    for url in URLS_NJ:
        NJ(driver, url)
        
    expireddomains2(date_exp_3_days, exp_date3)
    expireddomains2(date_exp_4_days, exp_date4) 
    expireddomains2(date_exp_5_days, exp_date5)
    
    excel_file1 = f"C:\\DomCop\\Name\\Name_NJ_{exp_date3}.xlsx"
    excel_file2 = f"C:\\DomCop\\Name\\Name_NJ_{exp_date4}.xlsx"
    excel_file3 = f"C:\\DomCop\\Name\\Name_NJ_{exp_date5}.xlsx"
    file_type = 'NJ'
    paste_data(file_type, exp_date3, excel_file1)
    paste_data(file_type, exp_date4, excel_file2)
    paste_data(file_type, exp_date5, excel_file3)
        
    for url in URLS_A2:
        A2(driver, url)
    
    expireddomains3(date_exp_3_days, exp_date3)
    expireddomains3(date_exp_4_days, exp_date4) 
    expireddomains3(date_exp_5_days, exp_date5)
    
    excel_file1 = f"C:\\DomCop\\Name\\Name_Auctions2_{exp_date3}.xlsx"
    excel_file2 = f"C:\\DomCop\\Name\\Name_Auctions2_{exp_date4}.xlsx"
    excel_file3 = f"C:\\DomCop\\Name\\Name_Auctions2_{exp_date5}.xlsx"
    file_type = 'Auctions2'
    paste_data(file_type, exp_date3, excel_file1)
    paste_data(file_type, exp_date4, excel_file2)
    paste_data(file_type, exp_date5, excel_file3)

driver.quit()