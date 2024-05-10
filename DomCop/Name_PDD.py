import os
import glob
import shutil
import config
import ftplib
import time
import openpyxl
import random
import pandas as pd
import zipfile
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

URLS = ["https://www.domcop.com/domains?sid=15583",
    "https://www.domcop.com/domains?sid=15649",
    "https://www.domcop.com/domains?sid=19380",
    "https://www.domcop.com/domains?sid=17308"]

service = webdriver.ChromeService()
options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')
options.add_argument("--disable-dev-shm-usage")
options.add_argument('log-level=3')
options.add_argument(r'user-data-dir=C:\Users\User3\AppData\Local\Google\Chrome\User Data')
prefs = {"download.default_directory" : r"C:\DomCop\Name\Files_PDD"}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(service=service, options=options)
print('.: NAME_PDD.BAT :.')

formatted_date = config.traffic_pdd_days_used.strftime("%d-%m-%Y")
exp_date = config.exp_date_today.strftime("%d-%m-%Y")


def extract_zip(zip_file, destination_path=None):
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(destination_path)
        print(f"Extracted '{zip_file}' to {destination_path or 'current directory'}")

def random_sleep(x, y):
    sleep_time = random.uniform(x, y)
    time.sleep(sleep_time)


def main(driver, url):
    try:
        #access the web page
        driver.get(url)
        random_sleep(10, 15)
        print("accessing web")
        driver.find_element(By.XPATH, '//button[@class="btn btn-primary btn-minier"][3]').click()
        print("export..")
        random_sleep(1, 2)
        driver.find_element(By.LINK_TEXT, 'Select None').click()
        random_sleep(3, 5)
        print("choosing menu..")
        driver.find_element(By.XPATH, '//input[@name="DomainData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//button[@class="btn btn-info"][1]').click()
        random_sleep(2, 5)
        print("waiting to download file...")

        try:
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")
        except:
            print('loading more... wait!')
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")

        # link = None
        # while not link:
        #     try:
        #         time.sleep(20)
        #         link = driver.find_element(By.XPATH, '//tr[2]/td[8]/a[@href]').click()
        #         link = 'done'
        #         print("download success...")
        #     except NoSuchElementException:
        #         time.sleep(30)        
        random_sleep(2, 5)
        # Renaming the and split file
        # download_path = glob.glob(r"C:\DomCop\Name\Files\*") 
        # latest_filename = max(download_path, key=os.path.getctime)
        # output_prefix = latest_filename[:-4]

    except Exception as error:
        print(error)
        print('something wrong.. browser quit')
        driver.quit()

def expireddomains():
    expireddomainsurl = 'https://member.expireddomains.net/savedsearches/goto/429442/'
    driver.get(expireddomainsurl)
    print('opening expireddomain..')
    random_sleep(5, 10)

    domain_data_list = []  # List to store extracted domain data

    try:
        # Find the "Show Filter" button (assuming it's always present)
        show_filter_button = driver.find_element(By.LINK_TEXT, "Show Filter")
        print('show filter..')

        # Click the button
        show_filter_button.click()
        random_sleep(1, 3)

        # Calculated date
        date_exp = config.traffic_pdd_days_used.strftime("%Y-%m-%d")

        # Select the date by ID
        select = Select(driver.find_element(By.ID, 'fenddate'))
        select.select_by_value(date_exp)
        print('choose date..')
        random_sleep(1, 3)

        # Find the "Submit" button by name
        button = driver.find_element(By.NAME, "button_submit")
        
        # Click the "Submit" button and wait
        button.click()
        print('apply filter..')
        random_sleep(3, 10)

        clickable = True
            
        while clickable == True:
            # Find domain data elements
            span_element = driver.find_element(By.CSS_SELECTOR, 'span[data-clipboard-text]')
            data_text = span_element.get_attribute('data-clipboard-text')
            domain_data_list.append(data_text)
            print('domain copied into list..')
            random_sleep(3, 5)
            
            next_button = ''
            try:
                #Check for "Next" button and click if available
                next_buttons = driver.find_elements(By.CLASS_NAME, "next")
                next_button = next_buttons[0]
                if next_button:
                    next_button.click()
                    print('next page..')
                    random_sleep(10, 15)
            except:
                clickable = False
                # No "Next" button found, break the loop
                break
               
        # Create the filename and folder structure
        folder_path = r"C:\DomCop\Name\Files_PDD"
        filename = f"Name_PDD_Expireddomains_{formatted_date}.txt"
        os.makedirs(folder_path, exist_ok=True)

        # Construct the full filepath
        filepath = os.path.join(folder_path, filename)

        # Open the file in write mode and write all extracted data at once
        with open(filepath, "w") as text_file:
            text_file.write("\n".join(domain_data_list))  # Efficient single write

        # Go to Google (optional)
        driver.get('https://google.com')

    except Exception as e:
        print(f"Something went wrong: {e}")
        driver.quit()

  
def copy_template():
    original_file = f'C:\DomCop\Template_Name_DomCop.xlsx'
    destination_directory = r'C:\DomCop\Name'
    
    # Construct the new filename
    new_filename = f"Name_PDD_{formatted_date}.xlsx"
    destination_path = os.path.join(destination_directory, new_filename)
    # Copy the file
    shutil.copyfile(original_file, destination_path)

    print(f"File duplicated as: {destination_path}")
    

def paste_data(excel_file):

    domain_list_file = f"C:\\DomCop\\Name\\Files_PDD\\Name_PDD_Expireddomains_{formatted_date}.txt"

    # Define your Excel file path

    # Open the text file and read the domain list
    with open(domain_list_file, "r") as f:
        domain_list = f.readlines()

    # Clean the domain list (optional, remove leading/trailing whitespaces)
    domain_list = [domain.strip() for domain in domain_list]
    
    # Define the directory containing CSV files
    csv_dir = r'C:\DomCop\Name\Files_PDD'  # Replace with the actual directory path

    # Find the latest 4 CSV files (modify as needed)
    # Find all CSV files in the directory
    csv_files = [os.path.join(csv_dir, filename) for filename in os.listdir(csv_dir) if filename.endswith(".csv")]
    
    random_sleep(1, 3)
    
    # Sort the CSV files by modification time (latest first)
    csv_files.sort(key=os.path.getmtime, reverse=True)

    # Limit the number of files to process (e.g., 4 or 5)
    num_files = min(4, len(csv_files))  # Process up to 4 or all available files (whichever is less)
    csv_files = csv_files[:num_files]
    
    # Read each CSV file and extract "Domain Name" column
    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file)
            domain_list.extend(df["Domain Name"].tolist())  # Efficiently append domain names
        except FileNotFoundError:
            print(f"Error: CSV file '{csv_file}' not found. Skipping...")
            
    # Lowercase all characters in the list
    domain_list = [domain.lower() for domain in domain_list]
    
    # Remove duplicates from domain_list using set conversion
    domain_list = list(set(domain_list))

    # Load the existing Excel file
    workbook = openpyxl.load_workbook(excel_file)

    # Check if the "filter" sheet exists
    if "filter" not in workbook.sheetnames:
        # Create the "filter" sheet at the beginning
        worksheet = workbook.create_sheet("filter", 0)
    else:
        # Get the existing "filter" sheet
        worksheet = workbook[ "filter"]

    # Clear formatting on the entire sheet
    for row in worksheet.iter_rows():
      row[1].alignment = None  # Clear alignment for the first cell (assuming data starts from A2)
      worksheet.row_dimensions[row[1].row].alignment = None 

    # Write the domain list starting from the first row (checks for existing data)
    if worksheet["A1"].value is None:
        row_counter = 1  # Start writing from the first row if empty
    else:
        row_counter = worksheet.max_row + 1  # Start after existing data

    for domain in domain_list:
        worksheet.cell(row=row_counter, column=1, value=domain)
        row_counter += 1

    # Save the Excel file
    workbook.save(excel_file)
    print("Domain list pasted successfully!")
    
if __name__ == "__main__":
    for url in URLS:
        main(driver, url)
        
    print("all CSV download success...") 
    expireddomains()
    print("copy template..")
    copy_template()
    print("paste data..")
    excel_file = f"C:\\DomCop\\Name\\Name_PDD_{formatted_date}.xlsx"
    paste_data(excel_file)
    
driver.quit()