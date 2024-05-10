import os
import glob
import shutil
import config
import ftplib
import time
import openpyxl
import random
import pandas as pd
from twocaptcha import TwoCaptcha
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

URLS = ["https://www.domcop.com/domains?sid=18642",
        "https://www.domcop.com/domains?sid=19379"]

service = webdriver.ChromeService()
options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')
options.add_argument("--disable-dev-shm-usage")
options.add_argument('log-level=3')
options.add_argument(r'user-data-dir=C:\Users\User3\AppData\Local\Google\Chrome\User Data')
prefs = {"download.default_directory" : r"C:\DomCop\Traffic"}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(service=service, options=options)
print('.: TRAFFIC.BAT :.')

def captcha_solver():
    print('captcha solver..')
    solver = TwoCaptcha("2527582599cc0c9d5c6645f3c6cbe63c")
    response = solver.recaptcha(sitekey='6LfLwI4UAAAAADRoaQZs3hBl9xzsmwqz126iqYEq', url=url)
    code = response['code']
    

def random_sleep(x, y):
    sleep_time = random.uniform(x, y)
    time.sleep(sleep_time)
    
def is_login():
    try:
        random_sleep(1,3)
        logout_icon = driver.find_element(By.XPATH, "//i[@class='icon-off']")
        print("login ok..")
        return True
    except:
        print("Element not found. Possibly logged out.")
        random_sleep(1,3)
        login_check()
    
def login_check():
    while True:
        url = 'https://www.domcop.com/login'
        driver.get(url)
        login_button = driver.find_element(By.ID, "button-submit-login-form")
        random_sleep(1,5)
        login_button.click()
        random_sleep(10,15)
        login_page_title = 'Login to your DomCop Account - DomCop'
        page_title = driver.title
        if page_title == login_page_title:
            print('possibly captcha..')
            captcha_solver()
            break
        else:
            break
               
def is_login_page():
    login_page_title = 'Login to your DomCop Account - DomCop'
    page_title = driver.title
    
    if page_title == login_page_title:
        print('on the login page..')
        login_check()
    else:
        pass
    

def filter_and_export_good_domains(csv_file, excel_file):
    # *** Filtering Logic: ***    
    # Load your Excel file, specifying the sheet name
    df = pd.read_csv(csv_file, header=0, index_col=False)

    # GOOD Filtering
    good_df = df[
        (df['Domain Age (WhoIs)'] == 0) | 
        (abs(df['Domain Age (WB)'] - df['Domain Age (WhoIs)']) < 4.0)
    ]

    # BAD Filtering (Remove from the 'good' results)
    bad_df = good_df[
        (good_df['Moz Spam Score'] >= 80) |
        ((good_df['Domain Age (WhoIs)'] == 0) & (good_df['Has Digits'] == 'Yes')) |
        ((good_df['Domain Age (WhoIs)'] == 0) & (good_df['Wayback Archive Crawls'] < 20)) |
        ((good_df['Domain Age (WhoIs)'] == 0) & (good_df['Majestic Ref Domains'] > 90)) |
        ((good_df['Domain Age (WhoIs)'] == 0) & (good_df['Moz DA 3yr%'] > 99)) 
    ]

    # Filter out BAD domains by removing the 'bad_df' rows from 'good_df'
    filtered_df = good_df[~good_df.index.isin(bad_df.index)]
    combined_df = filtered_df

    # *** Export Logic: ***
    # Load existing Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet_names = workbook.sheetnames

    # Check if "CSV" sheet exists, replace it completely
    if 'CSV' in sheet_names:
        idx = sheet_names.index('CSV')
        worksheet = workbook[sheet_names[idx]]  # Get existing "CSV" sheet
        worksheet.title = 'CSV'  # Rename the existing sheet to "CSV"

        # Clear existing sheet data (optional, comment out if not needed)
        # for row in worksheet.iter_rows():
        #    for cell in row:
        #        cell.value = None
    else:
        # Create "CSV" sheet at the desired position (e.g., index 0)
        worksheet = workbook.create_sheet('CSV', 0)

    # Write header
    for j, col_name in enumerate(combined_df.columns):
        worksheet.cell(row=1, column=j + 1, value=col_name)

    # Write data (start from the second row)
    for i, row in enumerate(combined_df.itertuples(index=False)):
        for j, value in enumerate(row):
            worksheet.cell(row=i + 2, column=j + 1, value=value)

    # Save the Excel file
    workbook.save(excel_file)
    print('Filtered data exported to Excel successfully!')
   
def main(driver, url):
    try:
        #access the web page
        driver.get(url)
        random_sleep(5, 10)
        print("accessing web")
        driver.find_element(By.XPATH, '//button[@class="btn btn-primary btn-minier"][3]').click()
        print("export..")
        random_sleep(1, 3)
        is_login_page()
        random_sleep(1, 3)
        driver.find_element(By.LINK_TEXT, 'Select None').click()
        random_sleep(3, 5)
        print("choosing menu..")
        driver.find_element(By.XPATH, '//input[@name="DomainData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//input[@name="SeomozData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//input[@name="MajesticData"]').click()
        random_sleep(1, 2)
        driver.find_element(By.XPATH, '//button[@class="btn btn-info"][1]').click()
        random_sleep(2, 5)
        print("waiting to download file...")

        try:
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")
        except:
            print('loading more... wait!')
            random_sleep(10, 15)
            link = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH,'//tr[2]/td[8]/a[@href]')))
            link.click()
            random_sleep(5, 10)
            print("download success...")

        # link = None
        # while not link:
        #     try:
        #         time.sleep(20)
        #         link = driver.find_element(By.XPATH, '//tr[2]/td[8]/a[@href]').click()
        #         link = 'done'
        #         print("download success...")
        #     except NoSuchElementException:
        #         time.sleep(30)        
        time.sleep(2)

        # Calculate the date 3 days from now
        formatted_date = config.traffic_pdd_days_used.strftime("%d-%m-%Y")
        gd_date = config.traffic_gd_days_used.strftime("%d-%m-%Y")
        
        # Renaming the file
        realpath = 'C:\DomCop\Traffic'
        download_path = glob.glob("C:\DomCop\Traffic\*") 
        time.sleep(2)  # Allow a short delay for download to complete
        latest_filename = max(download_path, key=os.path.getctime)
        print(f'Latest downloaded file is :{latest_filename}')
        
        desired_filename = ""  # Initialize empty, will set filename later
        if url == URLS[0]:
            desired_filename = f"Traffic_PDD_{formatted_date}.csv"
        elif url == URLS[1]:
            desired_filename = f"Traffic_GD_{gd_date}.csv"

        if desired_filename:
            src = os.path.join(realpath, latest_filename)
            dst = os.path.join(realpath, desired_filename)
            try:
                os.rename(src, dst)
            except Exception as e:
                print("create a duplicate name " + str(e))
                desired_filename = desired_filename[:-4] + "_1.csv" 
                dst = os.path.join(realpath, desired_filename)
                os.rename(src, dst)
            
            print("File renamed to:", desired_filename)
        else:
            print('rename failed..')
        

        print('create a Copy of template file...')
        original_file = f'C:\DomCop\Template_Traffic_DomCop.xlsx'
        destination_directory = f'C:\DomCop\Traffic'
        
        if url == URLS[0]:
            # Construct the new filename
            new_filename = f"Traffic_PDD_{formatted_date}.xlsx"
            destination_path = os.path.join(destination_directory, new_filename)
            # Copy the file
            shutil.copyfile(original_file, destination_path)

            print(f"File duplicated as: {destination_path}")
            time.sleep(1)
            
            print('copy csv into xlsx, only good domains...')
            csv_file = f"C:\DomCop\Traffic\Traffic_PDD_{formatted_date}.csv"
            excel_file = f"C:\DomCop\Traffic\Traffic_PDD_{formatted_date}.xlsx"
            filter_and_export_good_domains(csv_file, excel_file)
            time.sleep(1)
        elif url == URLS[1]:
            # Construct the new filename
            new_filename = f"Traffic_GD_{gd_date}.xlsx"
            destination_path = os.path.join(destination_directory, new_filename)
            # Copy the file
            shutil.copyfile(original_file, destination_path)

            print(f"File duplicated as: {destination_path}")
            time.sleep(1)
            
            print('copy csv into xlsx, only good domains...')
            csv_file = f"C:\DomCop\Traffic\Traffic_GD_{gd_date}.csv"
            excel_file = f"C:\DomCop\Traffic\Traffic_GD_{gd_date}.xlsx"
            filter_and_export_good_domains(csv_file, excel_file)
            time.sleep(1)
        
        print('Task finish..')

    except Exception as error:
        print(error)
        print('something wrong.. browser quit')
        driver.quit()
        
if __name__ == "__main__":
    for url in URLS:
        main(driver, url)
       

driver.quit()